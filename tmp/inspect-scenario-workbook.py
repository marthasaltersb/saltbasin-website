import json, sys
from openpyxl import load_workbook

wb = load_workbook(sys.argv[1], read_only=True, data_only=True)
out = {}
for ws in wb.worksheets:
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows, ()))
    samples = [list(row) for _, row in zip(range(3), rows)]
    out[ws.title] = {"rows": ws.max_row, "columns": ws.max_column, "header": header, "samples": samples}
print(json.dumps(out, indent=2, default=str))
