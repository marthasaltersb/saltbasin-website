from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

NAVY = "1B2A3B"
CREAM = "F5F0E8"
GOLD = "C4843A"
GOLD_TINT = "F3E3CE"
TEAL = "4A7C8E"
RISK_RED = "B5433A"
RISK_RED_TINT = "F4DCD9"
GREEN = "6B8F71"
INPUT_BLUE = "0000FF"

_HEADER_FONT = Font(name="Arial", color=CREAM, bold=True, size=10)
_HEADER_FILL = PatternFill("solid", fgColor=NAVY)
_TITLE_FONT = Font(name="Arial", color=NAVY, bold=True, size=16)
_SECTION_FONT = Font(name="Arial", color=GOLD, bold=True, size=12)
_BODY_FONT = Font(name="Arial", size=10)
_WRAP = Alignment(wrap_text=True, vertical="top")

_CONFIDENCE_FILL = {
    "HIGH": PatternFill("solid", fgColor=GREEN),
    "MEDIUM-HIGH": PatternFill("solid", fgColor=GOLD),
    "MEDIUM": PatternFill("solid", fgColor=GOLD),
    "LOW": PatternFill("solid", fgColor=RISK_RED),
}
_SEVERITY_FILL = {
    "critical": PatternFill("solid", fgColor=RISK_RED),
    "moderate": PatternFill("solid", fgColor=GOLD),
    "minor": PatternFill("solid", fgColor=CREAM),
}


def _header_row(ws, row_idx: int, headers: list[str], widths: list[int] | None = None):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _WRAP
        if widths:
            ws.column_dimensions[get_column_letter(col_idx)].width = widths[col_idx - 1]


def _write_row(ws, row_idx: int, values: list, fill: PatternFill | None = None):
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = _BODY_FONT
        cell.alignment = _WRAP
        if fill:
            cell.fill = fill


def _title(ws, text: str, row: int = 1):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = _TITLE_FONT


def _review_banner(ws, row: int):
    cell = ws.cell(
        row=row,
        column=1,
        value="DRAFT — HUMAN REVIEW REQUIRED BEFORE CLIENT USE. Nothing in this "
        "workbook has been sent or published.",
    )
    cell.font = Font(name="Arial", color=RISK_RED, bold=True, italic=True, size=10)


def _build_executive_dashboard(wb: Workbook, deliverable: dict):
    ws = wb.active
    ws.title = "Executive Dashboard"
    _title(ws, "Salt Basin Net Works | HandoverOS GTM Deliverable")
    ws.cell(row=2, column=1, value=f"Topic: {deliverable['topic']}").font = _SECTION_FONT
    client_name = deliverable.get("engagement_client_name")
    if client_name:
        ws.cell(row=3, column=1, value=f"Client: {client_name}")
    _review_banner(ws, row=4)
    ws.cell(row=6, column=1, value="EXECUTIVE SUMMARY").font = _SECTION_FONT
    summary_cell = ws.cell(row=7, column=1, value=deliverable["executive_summary"])
    summary_cell.alignment = _WRAP
    summary_cell.font = _BODY_FONT
    ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=6)
    ws.row_dimensions[7].height = 90
    ws.column_dimensions["A"].width = 60


def _build_benchmark_master(wb: Workbook, deliverable: dict):
    ws = wb.create_sheet("Verified Source Detail")
    _title(ws, "Benchmark Master — Verified Primary Source Statistics")
    headers = ["Metric", "Value", "Source", "Year", "Sample Size", "URL", "Relevance", "Secondary?"]
    _header_row(ws, 3, headers, widths=[26, 16, 22, 10, 20, 30, 34, 12])
    row = 4
    secondary_fill = PatternFill("solid", fgColor=RISK_RED_TINT)
    for item in deliverable["benchmark_master"]:
        _write_row(
            ws,
            row,
            [
                item["metric"],
                item["value"],
                item["source"],
                item["year"],
                item["sample_size"],
                item["url"],
                item["relevance_note"],
                "SECONDARY" if item["is_secondary_source"] else "",
            ],
            fill=secondary_fill if item["is_secondary_source"] else None,
        )
        row += 1


def _build_industry_breakdown(wb: Workbook, deliverable: dict):
    rows = deliverable.get("industry_breakdown") or []
    if not rows:
        return
    ws = wb.create_sheet("Industry Breakdown")
    _title(ws, "Benchmark by Industry")
    headers = [
        "Industry",
        "Observed Mechanism",
        "Root Cause",
        "Rate Estimate",
        "Source",
        "Program Resolution",
        "Q2R Stage Affected",
    ]
    _header_row(ws, 3, headers, widths=[18, 34, 30, 18, 20, 34, 22])
    row = 4
    for item in rows:
        _write_row(
            ws,
            row,
            [
                item["industry"],
                item["leakage_or_risk_mechanism"],
                item["root_cause"],
                item["rate_estimate"],
                item["rate_source"],
                item["program_resolution"],
                item["q2r_stage_affected"],
            ],
        )
        row += 1


def _build_assumptions_methodology(wb: Workbook, deliverable: dict):
    ws = wb.create_sheet("Assumptions & Methodology")
    _title(ws, "Assumptions & Methodology Register")
    am = deliverable["assumptions_methodology"]
    row = 3

    ws.cell(row=row, column=1, value="SECTION 1 — Verified Primary-Source Statistics (Used As-Is)").font = _SECTION_FONT
    row += 1
    headers1 = ["Statistic", "Value Used", "Primary Source", "Publication Date", "Sample Size", "URL", "How Applied"]
    _header_row(ws, row, headers1, widths=[24, 16, 22, 16, 20, 30, 30])
    row += 1
    for item in am["verified_statistics"]:
        _write_row(
            ws,
            row,
            [
                item["statistic_name"],
                item["value_used"],
                item["primary_source"],
                item["publication_date"],
                item["sample_size"],
                item["url"],
                item["how_applied"],
            ],
        )
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="SECTION 2 — Modeled Assumptions (Disclose in Presentations)").font = _SECTION_FONT
    row += 1
    headers2 = ["Assumption", "Value Used", "Conservative", "Base", "Optimistic", "Rationale", "Recommendation"]
    _header_row(ws, row, headers2, widths=[24, 14, 12, 12, 12, 34, 30])
    row += 1
    gold_tint_fill = PatternFill("solid", fgColor=GOLD_TINT)
    for item in am["modeled_assumptions"]:
        _write_row(
            ws,
            row,
            [
                item["assumption_name"],
                item["value_used"],
                item["conservative"],
                item["base"],
                item["optimistic"],
                item["rationale"],
                item["recommendation"],
            ],
            fill=gold_tint_fill,
        )
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="SECTION 3 — Scenario-to-Source Mapping").font = _SECTION_FONT
    row += 1
    headers3 = ["Scenario", "Mapped Source Category", "Direct Citation", "Inference / Gap", "Confidence", "Note"]
    _header_row(ws, row, headers3, widths=[26, 26, 30, 30, 14, 30])
    row += 1
    for item in am["scenario_source_mapping"]:
        start_row = row
        _write_row(
            ws,
            row,
            [
                item["scenario"],
                item["mapped_source_category"],
                item["direct_citation"],
                item["inference_gap"],
                item["confidence_level"],
                item["note"],
            ],
        )
        conf_fill = _CONFIDENCE_FILL.get(item["confidence_level"])
        if conf_fill:
            ws.cell(row=start_row, column=5).fill = conf_fill
        row += 1


def _build_impact_quantification(wb: Workbook, deliverable: dict, client_arr: float | None):
    ws = wb.create_sheet("Impact Quantification")
    iq = deliverable["impact_quantification"]
    _title(ws, "Impact Quantification — ROI Calculator")
    ws.cell(row=2, column=1, value="Enter ARR in the blue cell. Every other value recalculates live.")

    arr_row = 4
    ws.cell(row=arr_row, column=1, value="Annual Recurring Revenue (ARR)")
    arr_cell = ws.cell(row=arr_row, column=2, value=client_arr if client_arr is not None else 0)
    arr_cell.font = Font(name="Arial", color=INPUT_BLUE, bold=True, size=10)
    arr_ref = f"$B${arr_row}"

    recovery_row = 5
    ws.cell(row=recovery_row, column=1, value="Recovery Rate")
    ws.cell(row=recovery_row, column=2, value=iq["recovery_rate_pct"] / 100)
    ws.cell(row=recovery_row, column=2).number_format = "0.0%"
    ws.cell(row=recovery_row, column=3, value=iq["recovery_rate_source_note"])
    recovery_ref = f"$B${recovery_row}"

    header_row = 7
    headers = ["Scenario", "Conservative", "Base", "High", "3-Yr Cumulative (Base)", "Recovery Value (3-Yr)", "Confidence"]
    _header_row(ws, header_row, headers, widths=[30, 14, 14, 14, 18, 18, 14])

    row = header_row + 1
    first_data_row = row
    for scenario in iq["scenarios"]:
        ws.cell(row=row, column=1, value=scenario["scenario"])
        ws.cell(row=row, column=2, value=f"={scenario['conservative_rate_pct']/100}*{arr_ref}")
        ws.cell(row=row, column=3, value=f"={scenario['base_rate_pct']/100}*{arr_ref}")
        ws.cell(row=row, column=4, value=f"={scenario['high_rate_pct']/100}*{arr_ref}")
        ws.cell(row=row, column=5, value=f"=C{row}*3")
        ws.cell(row=row, column=6, value=f"=E{row}*{recovery_ref}")
        conf_cell = ws.cell(row=row, column=7, value=scenario["confidence_level"])
        conf_fill = _CONFIDENCE_FILL.get(scenario["confidence_level"])
        if conf_fill:
            conf_cell.fill = conf_fill
        for col in range(2, 7):
            ws.cell(row=row, column=col).number_format = "$#,##0;($#,##0);-"
            ws.cell(row=row, column=col).font = _BODY_FONT
        row += 1
    last_data_row = row - 1

    total_row = row
    ws.cell(row=total_row, column=1, value="TOTAL PORTFOLIO EXPOSURE").font = Font(name="Arial", bold=True, size=10)
    for col in range(2, 7):
        letter = get_column_letter(col)
        cell = ws.cell(row=total_row, column=col, value=f"=SUM({letter}{first_data_row}:{letter}{last_data_row})")
        cell.number_format = "$#,##0;($#,##0);-"
        cell.font = Font(name="Arial", bold=True, size=10)

    roi_row = total_row + 2
    ws.cell(row=roi_row, column=1, value="ROI SUMMARY").font = _SECTION_FONT
    ws.cell(row=roi_row + 1, column=1, value="Total 3-Year Recovery Value")
    ws.cell(row=roi_row + 1, column=2, value=f"=F{total_row}")
    ws.cell(row=roi_row + 1, column=2).number_format = "$#,##0;($#,##0);-"

    program_cost = iq.get("program_three_year_cost_usd")
    if program_cost is not None:
        ws.cell(row=roi_row + 2, column=1, value="Program Cost (3 Yrs)")
        ws.cell(row=roi_row + 2, column=2, value=program_cost)
        ws.cell(row=roi_row + 2, column=2).number_format = "$#,##0;($#,##0);-"
        ws.cell(row=roi_row + 3, column=1, value="Net ROI")
        ws.cell(row=roi_row + 3, column=2, value=f"=B{roi_row + 1}-B{roi_row + 2}")
        ws.cell(row=roi_row + 3, column=2).number_format = "$#,##0;($#,##0);-"

    multiple = iq.get("valuation_multiple")
    if multiple is not None:
        val_row = roi_row + 4 if program_cost is not None else roi_row + 2
        ws.cell(row=val_row, column=1, value=f"Valuation Impact at {multiple}x EV")
        ws.cell(row=val_row, column=2, value=f"=(B{roi_row + 1}/3)*{multiple}")
        ws.cell(row=val_row, column=2).number_format = "$#,##0;($#,##0);-"
        ws.cell(row=val_row, column=3, value=iq.get("valuation_multiple_source_note") or "")

    ws.column_dimensions["A"].width = 30


def _build_client_mapping(wb: Workbook, deliverable: dict):
    mapping = deliverable.get("client_mapping")
    if not mapping:
        return
    ws = wb.create_sheet("Client Actuals vs Benchmark")
    _title(ws, f"Client Data Mapping — {mapping['client_name']}")
    ws.cell(row=2, column=1, value=f"Target schema: {mapping['target_schema']}")

    row = 4
    ws.cell(row=row, column=1, value="FIELD MAPPING").font = _SECTION_FONT
    row += 1
    _header_row(ws, row, ["Raw Column", "Mapped Field", "Status", "Note"], widths=[26, 26, 14, 40])
    row += 1
    status_fill = {
        "confident": PatternFill("solid", fgColor=GREEN),
        "uncertain": PatternFill("solid", fgColor=GOLD),
        "unmapped": PatternFill("solid", fgColor=RISK_RED),
    }
    for item in mapping["field_mappings"]:
        _write_row(
            ws,
            row,
            [item["raw_column"], item["mapped_field"] or "—", item["mapping_status"], item["note"]],
        )
        fill = status_fill.get(item["mapping_status"])
        if fill:
            ws.cell(row=row, column=3).fill = fill
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="CLIENT ACTUALS VS. BENCHMARK").font = _SECTION_FONT
    row += 1
    _header_row(ws, row, ["Metric", "Client Value", "Benchmark Value", "Delta", "Confidence"], widths=[26, 18, 20, 34, 14])
    row += 1
    for item in mapping["client_actuals_vs_benchmark"]:
        _write_row(
            ws,
            row,
            [
                item["metric"],
                item["client_value"] if item["client_value"] is not None else "—",
                item["benchmark_value"],
                item["delta_description"],
                item["confidence_level"],
            ],
        )
        conf_fill = _CONFIDENCE_FILL.get(item["confidence_level"])
        if conf_fill:
            ws.cell(row=row, column=5).fill = conf_fill
        row += 1


def _build_data_quality_gaps(wb: Workbook, deliverable: dict):
    gaps = deliverable.get("data_quality_gaps") or []
    if not gaps:
        return
    ws = wb.create_sheet("Data Quality Gaps")
    _title(ws, "Data Quality Gaps")
    _header_row(ws, 3, ["Description", "Severity"], widths=[60, 14])
    row = 4
    for item in gaps:
        _write_row(ws, row, [item["description"], item["severity"]])
        fill = _SEVERITY_FILL.get(item["severity"])
        if fill:
            ws.cell(row=row, column=2).fill = fill
        row += 1


def _build_unverified_flags(wb: Workbook, deliverable: dict):
    flags = deliverable.get("unverified_flags") or []
    if not flags:
        return
    ws = wb.create_sheet("Unverified — Removed")
    _title(ws, "Unverified Claims — Excluded From Deliverable")
    ws.cell(row=2, column=1, value="These did not verify to a primary source and were left out of every other tab.")
    _header_row(ws, 4, ["Claim", "Reason Unverified"], widths=[50, 50])
    row = 5
    red_fill = PatternFill("solid", fgColor=RISK_RED_TINT)
    for item in flags:
        _write_row(ws, row, [item["claim"], item["reason_unverified"]], fill=red_fill)
        row += 1


def build_xlsx(deliverable: dict, output_path: Path, client_arr: float | None = None) -> None:
    wb = Workbook()
    _build_executive_dashboard(wb, deliverable)
    _build_benchmark_master(wb, deliverable)
    _build_industry_breakdown(wb, deliverable)
    _build_assumptions_methodology(wb, deliverable)
    _build_impact_quantification(wb, deliverable, client_arr)
    _build_client_mapping(wb, deliverable)
    _build_data_quality_gaps(wb, deliverable)
    _build_unverified_flags(wb, deliverable)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
